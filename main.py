from openpyxl import Workbook, load_workbook
import os

from clientes import cadastrar_clientes, listar_clientes, buscar_cliente, fechar, deletar_cliente, linha, pausar, editar_cliente

arquivo = "data/tabelaclientes.xlsx"

if os.path.exists(arquivo):
    wb = load_workbook(arquivo)
    ws = wb.active
else:
    wb = Workbook()
    ws = wb.active
    ws.append(["Nome", "Email", "Telefone"])

while True:
    while True:
        opcao = input(
"""MENU
1 - Cadastrar clientes.
2 - Lista de clientes.
3 - Buscar clientes.
4 - Deletar cliente.
5 - Editar cliente.
6 - Sair.
Escolha uma das opções: """
        )
        if opcao in ['1', '2', '3', '4', '5', '6']:
            break
        print("OPÇÃO INVÁLIDA.")
        linha()

    # CADASTRAR CLIENTES
    if opcao == '1':
        cadastrar_clientes(ws, wb, arquivo)
        pausar()


    # LISTA DE CLIENTES
    elif opcao == '2':
        listar_clientes(ws)
        pausar()

    # PESQUISAR CLIENTE
    elif opcao == '3':
        buscar_cliente(ws)
        pausar()

    # FINALIZAR
    elif opcao == '4':
        deletar_cliente(ws, wb, arquivo)
        pausar()

    elif opcao == '5':
        editar_cliente(ws, wb, arquivo)
        pausar()

    elif opcao == '6':
        fechar(wb, arquivo)
        break
